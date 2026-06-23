"""
Excel 通用生成器（从零生成模式）

配置驱动的 Excel 文档生成器，与模板填充模式(excel-testcase-generator.js)并存。
所有列定义、样式、数据源映射均从 config/excel-doc-types.json 读取，零硬编码。

架构分层：
  1. 配置加载层 - ExcelDocTypeConfig: 加载 excel-doc-types.json
  2. 数据解析层 - parsers: 根据 source_type 解析 MD/JSON/CSV 为统一数据行
  3. 样式构建层 - StyleFactory: 从配置构建 openpyxl 样式对象
  4. Excel 生成层 - ExcelBuilder: 组装 Sheet、写入数据、应用样式、输出 xlsx

使用方式:
  python excel_generator.py --doc-type test-case-custom \
    --md-files file1.md file2.md \
    --module "模块名称" \
    --output output.xlsx
"""
import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# Windows 下强制 stdout/stderr 使用 UTF-8，避免中文乱码
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelDocTypeConfig:
    """配置加载层：从 excel-doc-types.json 读取文档类型定义"""

    def __init__(self, config_path):
        self.config_path = Path(config_path)
        if not self.config_path.exists():
            raise FileNotFoundError(f"配置文件不存在: {self.config_path}")
        with open(self.config_path, "r", encoding="utf-8") as f:
            self._raw = json.load(f)
        self.defaults = self._raw.get("defaults", {})
        self.doc_types = self._raw.get("doc_types", {})

    def get_doc_type(self, doc_type):
        if doc_type not in self.doc_types:
            available = ", ".join(self.doc_types.keys())
            raise ValueError(f"不支持的文档类型: {doc_type}，可用: {available}")
        return self.doc_types[doc_type]

    def get_merged_styles(self, doc_type):
        """合并默认样式与文档类型级别覆盖（文档类型级优先）"""
        merged = json.loads(json.dumps(self.defaults.get("styles", {})))
        doc_styles = self.get_doc_type(doc_type).get("styles", {})
        for key, val in doc_styles.items():
            if isinstance(val, dict) and key in merged and isinstance(merged[key], dict):
                merged[key].update(val)
            else:
                merged[key] = val
        return merged

    def resolve_output_filename(self, doc_type, module, output_path=None):
        """解析输出文件名，支持 {module}/{docLabel}/{date} 变量"""
        if output_path:
            return output_path
        doc_def = self.get_doc_type(doc_type)
        pattern = self.defaults.get("output", {}).get(
            "filename_pattern", "{module}-{docLabel}-{date}.xlsx"
        )
        date_str = datetime.now().strftime("%Y%m%d")
        filename = pattern.format(
            module=module or "未命名",
            docLabel=doc_def.get("doc_label", "文档"),
            date=date_str,
        )
        # 默认输出到技能的 output 目录
        output_dir = self.config_path.parent.parent / "output"
        return str(output_dir / filename)


class StyleFactory:
    """样式构建层：从配置构建 openpyxl 样式对象，缓存复用"""

    def __init__(self, styles_config):
        self.cfg = styles_config
        self._cache = {}

    def _hex_to_color(self, hex_color):
        """规范化颜色值，openpyxl 需要 8 位 ARGB 或 6 位 RGB"""
        c = hex_color.lstrip("#").upper()
        return c if len(c) == 6 else c[-6:]

    def header_font(self):
        key = "header_font"
        if key not in self._cache:
            h = self.cfg.get("header", {})
            self._cache[key] = Font(
                bold=h.get("bold", True),
                color=self._hex_to_color(h.get("font_color", "FFFFFF")),
                size=h.get("font_size", 12),
                name=h.get("font_name", "Arial"),
            )
        return self._cache[key]

    def header_fill(self):
        key = "header_fill"
        if key not in self._cache:
            h = self.cfg.get("header", {})
            self._cache[key] = PatternFill(
                fill_type="solid",
                fgColor=self._hex_to_color(h.get("fill_color", "1F4E78")),
            )
        return self._cache[key]

    def data_font(self):
        key = "data_font"
        if key not in self._cache:
            d = self.cfg.get("data", {})
            self._cache[key] = Font(
                size=d.get("font_size", 11),
                name=d.get("font_name", "Arial"),
            )
        return self._cache[key]

    def bold_data_font(self):
        key = "bold_data_font"
        if key not in self._cache:
            d = self.cfg.get("data", {})
            self._cache[key] = Font(
                size=d.get("font_size", 11),
                name=d.get("font_name", "Arial"),
                bold=True,
            )
        return self._cache[key]

    def zebra_fill(self, row_index):
        """根据行索引返回斑马纹填充（0-based）"""
        z = self.cfg.get("zebra", {})
        color = z.get("color1", "FFFFFF") if row_index % 2 == 0 else z.get("color2", "F7F9FC")
        return PatternFill(fill_type="solid", fgColor=self._hex_to_color(color))

    def border(self):
        key = "border"
        if key not in self._cache:
            b = self.cfg.get("border", {})
            side = Side(style=b.get("style", "thin"), color=self._hex_to_color(b.get("color", "D9DEE7")))
            self._cache[key] = Border(left=side, right=side, top=side, bottom=side)
        return self._cache[key]

    def header_alignment(self):
        key = "header_align"
        if key not in self._cache:
            a = self.cfg.get("alignment", {})
            self._cache[key] = Alignment(
                horizontal=a.get("header_horizontal", "center"),
                vertical=a.get("header_vertical", "center"),
                wrap_text=a.get("wrap_text", True),
            )
        return self._cache[key]

    def data_alignment(self, horizontal=None):
        """数据单元格对齐，允许列级别覆盖"""
        a = self.cfg.get("alignment", {})
        h = horizontal or a.get("data_horizontal_default", "left")
        return Alignment(
            horizontal=h,
            vertical=a.get("vertical", "center"),
            wrap_text=a.get("wrap_text", True),
        )

    def header_row_height(self):
        return self.cfg.get("row_height", {}).get("header", 30)

    def data_row_height(self):
        return self.cfg.get("row_height", {}).get("data", 80)


class TestCaseMdParser:
    """数据解析层：解析测试用例 MD 文件为统一数据行"""

    def __init__(self, field_mappings, validation_config=None):
        self.mappings = field_mappings
        self.validation_config = validation_config or {}

    def parse(self, md_files):
        """解析多个 MD 文件，返回数据行列表和校验报告"""
        all_rows = []
        validation_report = {"warnings": [], "errors": [], "skipped_files": []}

        for md_file in md_files:
            if not os.path.exists(md_file):
                msg = f"文件不存在，跳过: {md_file}"
                print(f"  ⚠ {msg}", file=sys.stderr)
                validation_report["skipped_files"].append(md_file)
                continue

            with open(md_file, "r", encoding="utf-8") as f:
                content = f.read()

            # MD 格式校验（解析前）
            fmt_check = self._validate_md_format(content, md_file)
            validation_report["warnings"].extend(fmt_check["warnings"])
            validation_report["errors"].extend(fmt_check["errors"])

            if fmt_check["errors"]:
                print(f"  ⚠ {md_file} 格式校验未通过，仍尝试解析", file=sys.stderr)

            rows = self._parse_content(content, md_file)
            all_rows.extend(rows)
            print(f"  已解析: {md_file} -> {len(rows)} 条记录")

        return all_rows, validation_report

    def _validate_md_format(self, content, file_path):
        """校验 MD 文件是否包含必要结构（基于 validation.md_format_check 配置）"""
        result = {"warnings": [], "errors": []}
        fmt_check = self.validation_config.get("md_format_check", {})

        must_contain = fmt_check.get("must_contain", [])
        for pattern in must_contain:
            if pattern not in content:
                result["errors"].append(f"{os.path.basename(file_path)}: 缺少必要内容 '{pattern}'")

        warn_if_missing = fmt_check.get("warn_if_missing", [])
        for pattern in warn_if_missing:
            if pattern not in content:
                result["warnings"].append(f"{os.path.basename(file_path)}: 建议包含 '{pattern}'")

        return result

    def _parse_content(self, content, file_path=""):
        """按 case_id_pattern 分割用例，逐个提取字段（单条容错）"""
        case_id_pattern = self.mappings.get("case_id_pattern", r"^## (TC-\S+)")
        parts = re.split(case_id_pattern, content, flags=re.MULTILINE)
        rows = []
        # parts[0] 是标题部分，之后每两个元素一组（编号+内容）
        for i in range(1, len(parts), 2):
            case_id = parts[i].strip()
            case_content = parts[i + 1] if i + 1 < len(parts) else ""
            try:
                row = self._extract_fields(case_id, case_content)
                rows.append(row)
            except Exception as e:
                print(f"  ⚠ 用例 {case_id} 解析失败: {e}，跳过", file=sys.stderr)
        return rows

    def _extract_fields(self, case_id, content):
        """根据 field_mappings 提取每个字段"""
        row = {}
        for field_name, mapping in self.mappings.get("field_mappings", {}).items():
            source = mapping.get("source", "field")
            if source == "case_id":
                row[field_name] = case_id
            elif source == "field":
                regex = mapping.get("regex")
                default = mapping.get("default", "-")
                if regex:
                    m = re.search(regex, content)
                    row[field_name] = m.group(1).strip() if m else default
                else:
                    row[field_name] = default
            elif source == "section":
                section = mapping.get("section")
                item_prefix = mapping.get("item_prefix", "- [.] ")
                row[field_name] = self._extract_section_items(content, section, item_prefix)
            elif source == "section_text":
                section = mapping.get("section")
                row[field_name] = self._extract_section_text(content, section)
            elif source == "table":
                section = mapping.get("section")
                fmt = mapping.get("format", "step")
                row[field_name] = self._extract_table(content, section, fmt)
            else:
                row[field_name] = mapping.get("default", "-")
        return row

    def _extract_section_items(self, content, section_name, item_prefix):
        """提取章节下的勾选项列表"""
        pattern = rf"### {re.escape(section_name)}\s*\n(.*?)(?=\n### |\Z)"
        m = re.search(pattern, content, re.DOTALL)
        if not m:
            return "-"
        text = m.group(1).strip()
        escaped_prefix = re.escape(item_prefix)
        items = re.findall(rf"{escaped_prefix}(.+)", text)
        return "\n".join(f"• {item}" for item in items) if items else text

    def _extract_section_text(self, content, section_name):
        """提取章节下的纯文本（到下一个 ### 或 ---）"""
        pattern = rf"### {re.escape(section_name)}\s*\n\n(.*?)(?=\n### |\n---|\Z)"
        m = re.search(pattern, content, re.DOTALL)
        return m.group(1).strip() if m else "-"

    def _extract_table(self, content, section_name, fmt):
        """提取章节下的表格数据"""
        pattern = rf"### {re.escape(section_name)}\s*\n\|.*?\|.*?\|\n\|[-|]+\|(.*?)(?=\n### |\Z)"
        m = re.search(pattern, content, re.DOTALL)
        if not m:
            return "-"
        table_text = m.group(1).strip()
        rows = re.findall(r"\|\s*(\d+)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|", table_text)
        if fmt == "step":
            lines = [f"步骤{r[0]}: {r[1]}\n  预期: {r[2]}" for r in rows]
            return "\n".join(lines)
        return "\n".join(f" | ".join(r) for r in rows)


class JsonDataParser:
    """数据解析层：解析 JSON 数据文件为统一数据行"""

    def __init__(self, field_mappings, validation_config=None):
        self.mappings = field_mappings
        self.validation_config = validation_config or {}

    def parse(self, json_files):
        all_rows = []
        validation_report = {"warnings": [], "errors": [], "skipped_files": []}

        for json_file in json_files:
            if not os.path.exists(json_file):
                print(f"  ⚠ 文件不存在，跳过: {json_file}", file=sys.stderr)
                validation_report["skipped_files"].append(json_file)
                continue
            with open(json_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                all_rows.extend(data)
            elif isinstance(data, dict) and "test_cases" in data:
                all_rows.extend(data["test_cases"])
            elif isinstance(data, dict):
                all_rows.append(data)
            print(f"  已解析: {json_file} -> {len(all_rows)} 条记录")

        return all_rows, validation_report


class ParserFactory:
    """解析器工厂：根据 source_type 返回对应解析器"""

    @staticmethod
    def create(data_source_config, validation_config=None):
        source_type = data_source_config.get("type", "md")
        parser_type = data_source_config.get("parser", "test_case_md")
        if source_type == "md" and parser_type == "test_case_md":
            return TestCaseMdParser(data_source_config, validation_config)
        elif source_type == "json":
            return JsonDataParser(data_source_config, validation_config)
        else:
            raise ValueError(f"不支持的解析器: source_type={source_type}, parser={parser_type}")


class DataValidator:
    """数据校验层：检查必填字段完整性，基于 validation.required_fields 配置"""

    def __init__(self, validation_config):
        self.required_fields = validation_config.get("required_fields", [])

    def validate(self, data_rows):
        """校验数据行中必填字段是否完整，返回校验报告"""
        report = {"missing_count": 0, "missing_details": []}

        if not self.required_fields:
            return report

        for idx, row in enumerate(data_rows):
            for field in self.required_fields:
                value = row.get(field, "")
                if not value or value == "-":
                    case_id = row.get("用例编号", f"行{idx + 1}")
                    report["missing_count"] += 1
                    report["missing_details"].append(
                        f"{case_id}: 缺少必填字段 '{field}'"
                    )

        return report


class ExcelBuilder:
    """Excel 生成层：组装 Sheet、写入数据、应用样式、输出 xlsx"""

    def __init__(self, doc_type_config, styles_config):
        self.doc_config = doc_type_config
        self.styles = StyleFactory(styles_config)

    def build(self, data_rows, output_path, module_name=""):
        """构建 Excel 文件"""
        wb = Workbook()
        # 移除默认 Sheet
        wb.remove(wb.active)

        sheets_config = self.doc_config.get("sheets", [])
        for sheet_cfg in sheets_config:
            self._write_data_sheet(wb, sheet_cfg, data_rows)

        # 汇总 Sheet
        summary_cfg = self.doc_config.get("summary_sheet", {})
        if summary_cfg.get("enabled", False):
            self._write_summary_sheet(wb, summary_cfg, data_rows)

        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
        wb.save(output_path)
        return output_path

    def _write_data_sheet(self, wb, sheet_cfg, data_rows):
        """写入数据 Sheet"""
        ws = wb.create_sheet(sheet_cfg["name"])
        columns = sheet_cfg.get("columns", [])

        # 写表头
        for col_idx, col_def in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["name"])
            cell.font = self.styles.header_font()
            cell.fill = self.styles.header_fill()
            cell.alignment = self.styles.header_alignment()
            cell.border = self.styles.border()
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get("width", 15)

        # 写数据
        for row_idx, row_data in enumerate(data_rows, 2):
            fill = self.styles.zebra_fill(row_idx - 2)
            for col_idx, col_def in enumerate(columns, 1):
                field_name = col_def["name"]
                value = row_data.get(field_name, "-")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles.data_font()
                cell.fill = fill
                cell.border = self.styles.border()
                align_h = col_def.get("align")
                cell.alignment = self.styles.data_alignment(align_h)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = self.styles.header_row_height()
        for row_idx in range(2, len(data_rows) + 2):
            ws.row_dimensions[row_idx].height = self.styles.data_row_height()

    def _write_summary_sheet(self, wb, summary_cfg, data_rows):
        """写入汇总 Sheet"""
        ws = wb.create_sheet(summary_cfg["name"])
        columns = summary_cfg.get("columns", [])
        group_field = summary_cfg.get("group_by_field")
        count_field = summary_cfg.get("count_field")
        priority_fields = summary_cfg.get("priority_fields", [])
        priority_source = summary_cfg.get("priority_field_source", "优先级")

        # 写表头
        for col_idx, col_def in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def["name"])
            cell.font = self.styles.header_font()
            cell.fill = self.styles.header_fill()
            cell.alignment = self.styles.header_alignment()
            cell.border = self.styles.border()
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get("width", 15)

        # 统计分组
        groups = {}
        for row in data_rows:
            group_val = row.get(group_field, "未分组")
            if group_val not in groups:
                groups[group_val] = {"count": 0, "priorities": {p: 0 for p in priority_fields}}
            groups[group_val]["count"] += 1
            pri_val = row.get(priority_source, "")
            if pri_val in groups[group_val]["priorities"]:
                groups[group_val]["priorities"][pri_val] += 1

        # 合计行
        total_count = 0
        total_priorities = {p: 0 for p in priority_fields}

        # 写数据
        row_idx = 2
        for group_val, stats in groups.items():
            fill = self.styles.zebra_fill(row_idx - 2)
            values = [group_val, stats["count"]] + [stats["priorities"][p] for p in priority_fields]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles.data_font()
                cell.fill = fill
                cell.border = self.styles.border()
                cell.alignment = self.styles.data_alignment("center")
            total_count += stats["count"]
            for p in priority_fields:
                total_priorities[p] += stats["priorities"][p]
            row_idx += 1

        # 合计行（加粗）
        fill = self.styles.zebra_fill(row_idx - 2)
        values = ["合计", total_count] + [total_priorities[p] for p in priority_fields]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = self.styles.bold_data_font()
            cell.fill = fill
            cell.border = self.styles.border()
            cell.alignment = self.styles.data_alignment("center")

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = self.styles.header_row_height()


def main():
    parser = argparse.ArgumentParser(description="配置驱动的 Excel 文档生成器")
    parser.add_argument("--doc-type", required=True, help="文档类型（见 excel-doc-types.json）")
    parser.add_argument("--md-files", nargs="+", help="MD 数据源文件路径（多个）")
    parser.add_argument("--json-files", nargs="+", help="JSON 数据源文件路径（多个）")
    parser.add_argument("--module", default="", help="模块名称（用于输出文件名）")
    parser.add_argument("--output", help="输出文件路径（不指定则按 pattern 生成）")
    parser.add_argument(
        "--config",
        default=os.path.join(os.path.dirname(__file__), "..", "..", "config", "excel-doc-types.json"),
        help="配置文件路径",
    )
    args = parser.parse_args()

    # 1. 加载配置
    config = ExcelDocTypeConfig(args.config)
    doc_def = config.get_doc_type(args.doc_type)
    print(f"文档类型: {args.doc_type} | {doc_def.get('description', '')}")

    # 2. 确定数据源文件
    data_source = doc_def.get("data_source", {})
    source_type = data_source.get("type", "md")
    if source_type == "md":
        source_files = args.md_files or []
    elif source_type == "json":
        source_files = args.json_files or []
    else:
        source_files = []

    sheets = doc_def.get("sheets", [])
    if sheets:
        sheet0 = sheets[0]
        required = sheet0.get("source_files_required", False)
        if required and not source_files:
            arg_name = sheet0.get("source_files_arg", "--md-files")
            print(f"❌ 错误: 必须提供数据源文件，使用 {arg_name} 参数", file=sys.stderr)
            sys.exit(1)

    if not source_files:
        print("❌ 错误: 未提供数据源文件", file=sys.stderr)
        sys.exit(1)

    # 3. 解析数据
    print(f"\n[1/4] 解析数据源 ({source_type})...")
    validation_config = doc_def.get("validation", {})
    parser_instance = ParserFactory.create(data_source, validation_config)
    data_rows, parse_report = parser_instance.parse(source_files)
    print(f"  共解析 {len(data_rows)} 条记录")

    # 输出解析阶段校验结果
    if parse_report["warnings"]:
        for w in parse_report["warnings"][:10]:
            print(f"  ⚠ {w}", file=sys.stderr)
    if parse_report["errors"]:
        for e in parse_report["errors"][:10]:
            print(f"  ⚠ {e}", file=sys.stderr)
    if parse_report["skipped_files"]:
        print(f"  ⚠ 跳过 {len(parse_report['skipped_files'])} 个文件", file=sys.stderr)

    if not data_rows:
        print("⚠ 警告: 未解析到任何数据，将生成空 Excel", file=sys.stderr)

    # 4. 必填字段校验
    print(f"\n[2/4] 必填字段校验...")
    validator = DataValidator(validation_config)
    field_report = validator.validate(data_rows)
    if field_report["missing_count"] > 0:
        print(f"  ⚠ {field_report['missing_count']} 个字段缺失", file=sys.stderr)
        for detail in field_report["missing_details"][:10]:
            print(f"    - {detail}", file=sys.stderr)
    else:
        print(f"  ✓ 必填字段校验通过")

    # 5. 生成 Excel
    print(f"\n[3/4] 生成 Excel...")
    styles = config.get_merged_styles(args.doc_type)
    builder = ExcelBuilder(doc_def, styles)
    output_path = config.resolve_output_filename(args.doc_type, args.module, args.output)
    builder.build(data_rows, output_path, args.module)
    print(f"  输出文件: {output_path}")

    # 6. 输出结果 JSON（供 cli.js 解析）
    print(f"\n[4/4] 完成")
    result = {
        "success": True,
        "docType": args.doc_type,
        "outputPath": output_path,
        "totalRecords": len(data_rows),
        "sheets": [s["name"] for s in doc_def.get("sheets", [])],
        "summarySheet": doc_def.get("summary_sheet", {}).get("enabled", False),
        "validation": {
            "parseWarnings": len(parse_report["warnings"]),
            "parseErrors": len(parse_report["errors"]),
            "skippedFiles": len(parse_report["skipped_files"]),
            "missingFields": field_report["missing_count"],
        },
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
