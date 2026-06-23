# -*- coding: utf-8 -*-
"""
SQL增量脚本变更分析器 - 核心脚本
自动解析DDL/DML SQL脚本，提取表结构变更信息，生成Excel汇总文档
"""
import os
import sys
import re
import json
import argparse
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ============================================================
# 配置加载
# ============================================================

def load_config(config_path=None):
    """加载配置文件，支持环境变量覆盖"""
    if config_path is None:
        config_path = os.path.join(os.path.dirname(__file__), '..', 'config', 'sql-change-config.json')
    config_path = os.path.normpath(config_path)

    if not os.path.exists(config_path):
        print(f"ERROR: Config file not found: {config_path}")
        sys.exit(1)

    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    # 环境变量覆盖项目根目录
    project_root = os.environ.get(config.get('project_root_env', 'BEMP_PROJECT_ROOT'), '.')
    config['_project_root'] = os.path.abspath(project_root)

    return config


def resolve_path(config, key):
    """将配置中的相对路径解析为绝对路径"""
    path = config.get(key, '')
    if os.path.isabs(path):
        return path
    return os.path.normpath(os.path.join(config['_project_root'], path))


# ============================================================
# SQL文件扫描
# ============================================================

def scan_sql_files(sql_dir, config):
    """扫描SQL目录，按DDL/DML分类"""
    ddl_suffix = config['sql_patterns']['ddl_file_suffix']
    dml_suffix = config['sql_patterns']['dml_file_suffix']

    ddl_files = []
    dml_files = []

    if not os.path.exists(sql_dir):
        print(f"ERROR: SQL directory not found: {sql_dir}")
        sys.exit(1)

    for fname in sorted(os.listdir(sql_dir)):
        fpath = os.path.join(sql_dir, fname)
        if not os.path.isfile(fpath):
            continue
        if fname.lower().endswith(ddl_suffix):
            ddl_files.append(fpath)
        elif fname.lower().endswith(dml_suffix):
            dml_files.append(fpath)

    print(f"Scanned: {len(ddl_files)} DDL files, {len(dml_files)} DML files")
    return ddl_files, dml_files


def read_sql_file(filepath, encoding='utf-8'):
    """读取SQL文件内容，处理BOM"""
    with open(filepath, 'r', encoding=encoding) as f:
        content = f.read()
    # 去除BOM
    if content.startswith('\ufeff'):
        content = content[1:]
    return content


# ============================================================
# DDL解析
# ============================================================

def parse_ddl_files(ddl_files, config):
    """解析所有DDL文件，返回字段变更和索引变更列表"""
    patterns = config['sql_patterns']
    op_map = config['operation_type_map']

    field_changes = []   # (table, field, comment, type_str, op_type)
    index_changes = []   # (table, index_name, index_fields, op_type)
    new_tables = {}      # table_name -> [(field, type_str, comment)]
    table_comments = {}  # table_name -> comment
    column_comments = {} # (table_lower, field_lower) -> comment

    for fpath in ddl_files:
        content = read_sql_file(fpath, config.get('encoding', 'utf-8'))
        fname = os.path.basename(fpath)
        # 统一换行，去除多余空白
        sql = re.sub(r'\r\n', '\n', content)

        # 1. 提取COMMENT ON COLUMN
        for m in re.finditer(patterns['comment_on_column'], sql, re.IGNORECASE):
            tbl = m.group(1).strip('"').lower()
            col = m.group(2).strip('"').lower()
            cmt = m.group(3).strip()
            column_comments[(tbl, col)] = cmt

        # 2. 提取COMMENT ON TABLE
        for m in re.finditer(patterns['comment_on_table'], sql, re.IGNORECASE):
            tbl = m.group(1).strip('"').lower()
            cmt = m.group(2).strip()
            table_comments[tbl] = cmt

        # 3. 解析CREATE TABLE
        for m in re.finditer(patterns['create_table'], sql, re.IGNORECASE):
            table_name = m.group(1).strip('"')
            table_lower = table_name.lower()
            # 提取CREATE TABLE后的括号内容
            start = m.end()
            # 找到匹配的右括号
            depth = 1
            pos = start
            while pos < len(sql) and depth > 0:
                if sql[pos] == '(':
                    depth += 1
                elif sql[pos] == ')':
                    depth -= 1
                pos += 1
            body = sql[start:pos-1]

            # 解析字段定义
            fields = parse_create_table_body(body, table_name, column_comments, config)
            new_tables[table_lower] = fields

        # 4. 解析ALTER TABLE ADD (多字段括号形式)
        for m in re.finditer(patterns['alter_add'], sql, re.IGNORECASE):
            table_name = m.group(1).strip('"')
            table_lower = table_name.lower()
            start = m.end()
            # 找到匹配的右括号
            depth = 1
            pos = start
            while pos < len(sql) and depth > 0:
                if sql[pos] == '(':
                    depth += 1
                elif sql[pos] == ')':
                    depth -= 1
                pos += 1
            body = sql[start:pos-1]
            fields = parse_alter_add_body(body, table_name, column_comments, config)
            for f in fields:
                field_changes.append((table_name, f[0], f[1], f[2], op_map['alter_add']))

        # 5. 解析ALTER TABLE ADD (单字段无括号形式)
        for m in re.finditer(patterns['alter_add_single'], sql, re.IGNORECASE):
            table_name = m.group(1).strip('"')
            field_name = m.group(2).strip('"')
            table_lower = table_name.lower()
            field_lower = field_name.lower()
            # 避免与括号形式重复
            already = any(fc[0].lower() == table_lower and fc[1].lower() == field_lower for fc in field_changes)
            if not already:
                # 提取类型：从字段名后到行尾或分号
                rest = sql[m.end():].split(';')[0].split('\n')[0].strip()
                type_str = rest.split()[0] if rest else ''
                type_str = normalize_type(type_str, rest, config)
                comment = column_comments.get((table_lower, field_lower), '')
                field_changes.append((table_name, field_name, comment, type_str, op_map['alter_add']))

        # 6. 解析ALTER TABLE MODIFY
        for m in re.finditer(patterns['alter_modify'], sql, re.IGNORECASE):
            table_name = m.group(1).strip('"')
            start = m.end()
            depth = 1
            pos = start
            while pos < len(sql) and depth > 0:
                if sql[pos] == '(':
                    depth += 1
                elif sql[pos] == ')':
                    depth -= 1
                pos += 1
            body = sql[start:pos-1]
            fields = parse_alter_add_body(body, table_name, column_comments, config)
            for f in fields:
                field_changes.append((table_name, f[0], f[1], f[2], op_map['alter_modify']))

        # 7. 解析ALTER TABLE DROP COLUMN
        for m in re.finditer(patterns['alter_drop_column'], sql, re.IGNORECASE):
            table_name = m.group(1).strip('"')
            field_name = m.group(2).strip('"')
            field_changes.append((table_name, field_name, '', '', op_map['alter_drop_column']))

        # 8. 解析CREATE INDEX
        for m in re.finditer(patterns['create_index'], sql, re.IGNORECASE):
            index_name = m.group(1).strip('"')
            table_name = m.group(2).strip('"')
            # 提取索引字段
            start = m.end()
            end = sql.find(')', start)
            idx_fields = sql[start:end].strip() if end > start else ''
            # 判断是新增表索引还是已有表索引
            table_lower = table_name.lower()
            if table_lower in new_tables:
                op_type = op_map.get('create_index_on_new_table', '新增表索引')
            else:
                op_type = op_map['create_index']
            index_changes.append((table_name, index_name, idx_fields, op_type))

        # 9. 解析DROP TABLE
        for m in re.finditer(patterns['drop_table'], sql, re.IGNORECASE):
            table_name = m.group(1).strip('"')
            field_changes.append((table_name, '', '', '', op_map['drop_table']))

        # 10. 解析DROP INDEX
        for m in re.finditer(patterns['drop_index'], sql, re.IGNORECASE):
            index_name = m.group(1).strip('"')
            index_changes.append(('', index_name, '', op_map['drop_index']))

    # 将新增表展开为逐字段记录
    for table_lower, fields in new_tables.items():
        # 找原始表名（保留大小写）
        table_name = table_lower
        for fc in field_changes:
            if fc[0].lower() == table_lower:
                table_name = fc[0]
                break
        for idx in index_changes:
            if idx[0].lower() == table_lower:
                table_name = idx[0]
                break
        for f in fields:
            field_changes.append((table_name, f[0], f[1], f[2], op_map['create_table']))

    return field_changes, index_changes


def parse_create_table_body(body, table_name, column_comments, config):
    """解析CREATE TABLE的字段定义体"""
    fields = []
    table_lower = table_name.lower()

    # 按逗号分割字段（需处理嵌套括号）
    parts = split_by_top_comma(body)

    for part in parts:
        part = part.strip()
        if not part:
            continue
        # 跳过约束行（PRIMARY KEY, CONSTRAINT, UNIQUE, CHECK, FOREIGN KEY）
        if re.match(r'(?i)^\s*(PRIMARY\s+KEY|CONSTRAINT|UNIQUE|CHECK|FOREIGN\s+KEY)', part):
            continue

        # 提取字段名和类型
        tokens = part.split()
        if len(tokens) < 2:
            continue

        field_name = tokens[0].strip('"').strip()
        field_lower = field_name.lower()

        # 跳过约束关键字误识别
        if field_name.upper() in ('PRIMARY', 'CONSTRAINT', 'UNIQUE', 'CHECK', 'FOREIGN'):
            continue

        # 提取类型
        type_str = normalize_type(tokens[1], part, config)
        comment = column_comments.get((table_lower, field_lower), '')

        fields.append((field_name, comment, type_str))

    return fields


def parse_alter_add_body(body, table_name, column_comments, config):
    """解析ALTER TABLE ADD的字段定义体"""
    return parse_create_table_body(body, table_name, column_comments, config)


def split_by_top_comma(text):
    """按顶层逗号分割，忽略括号内的逗号"""
    parts = []
    depth = 0
    current = []
    for ch in text:
        if ch == '(':
            depth += 1
            current.append(ch)
        elif ch == ')':
            depth -= 1
            current.append(ch)
        elif ch == ',' and depth == 0:
            parts.append(''.join(current))
            current = []
        else:
            current.append(ch)
    if current:
        parts.append(''.join(current))
    return parts


def normalize_type(type_token, full_def, config):
    """规范化字段类型为标准格式（如 varchar2(250)、number(11,6)）"""
    type_map = config.get('field_type_map', {})

    # type_token可能已包含括号（如 VARCHAR2(1)），先提取基础类型名（含尾部数字如VARCHAR2）
    base_match = re.match(r'([A-Za-z_]+\d*)', type_token)
    if not base_match:
        return type_token.lower()
    base_name = base_match.group(1).upper()
    mapped = type_map.get(base_name, base_name.lower())

    # 优先从type_token自身提取括号精度
    prec_match = re.search(r'\(([^)]+)\)', type_token)
    if prec_match:
        return f"{mapped}({prec_match.group(1)})"

    # 再从full_def提取括号精度
    prec_match = re.search(r'\(([^)]+)\)', full_def)
    if prec_match:
        return f"{mapped}({prec_match.group(1)})"

    return mapped


# ============================================================
# DML解析
# ============================================================

def parse_dml_files(dml_files, config):
    """解析所有DML文件，返回配置数据变更列表"""
    patterns = config['sql_patterns']
    op_map = config['operation_type_map']
    merge_rules = config.get('merge_rules', {})

    dml_changes = []  # (table, op_type, summary, id_key)

    for fpath in dml_files:
        content = read_sql_file(fpath, config.get('encoding', 'utf-8'))
        fname = os.path.basename(fpath)
        sql = re.sub(r'\r\n', '\n', content)

        # 提取文件名中的需求描述
        req_desc = extract_req_desc(fname)

        # 收集本文件中的DELETE和INSERT目标表
        deletes = []
        inserts = []
        updates = []

        for m in re.finditer(patterns['delete_from'], sql, re.IGNORECASE):
            tbl = m.group(1).strip('"')
            deletes.append(tbl.lower())

        for m in re.finditer(patterns['insert_into'], sql, re.IGNORECASE):
            tbl = m.group(1).strip('"')
            inserts.append(tbl.lower())

        for m in re.finditer(patterns['update_table'], sql, re.IGNORECASE):
            tbl = m.group(1).strip('"')
            updates.append(tbl.lower())

        # DELETE+INSERT合并逻辑
        if merge_rules.get('delete_insert_same_table', True):
            # 找出同时出现DELETE和INSERT的表
            merged_tables = set(deletes) & set(inserts)
            for tbl in merged_tables:
                id_keys = extract_id_keys(sql, tbl, 'INSERT')
                dml_changes.append((tbl, op_map.get('delete_insert', 'DELETE+INSERT'), req_desc, id_keys))
            # 仅DELETE的表
            for tbl in set(deletes) - merged_tables:
                id_keys = extract_id_keys(sql, tbl, 'DELETE')
                dml_changes.append((tbl, 'DELETE', req_desc, id_keys))
            # 仅INSERT的表
            for tbl in set(inserts) - merged_tables:
                id_keys = extract_id_keys(sql, tbl, 'INSERT')
                dml_changes.append((tbl, 'INSERT', req_desc, id_keys))
        else:
            for tbl in inserts:
                id_keys = extract_id_keys(sql, tbl, 'INSERT')
                dml_changes.append((tbl, 'INSERT', req_desc, id_keys))
            for tbl in deletes:
                id_keys = extract_id_keys(sql, tbl, 'DELETE')
                dml_changes.append((tbl, 'DELETE', req_desc, id_keys))

        # UPDATE
        for tbl in updates:
            id_keys = extract_id_keys(sql, tbl, 'UPDATE')
            dml_changes.append((tbl, op_map.get('update', 'UPDATE'), req_desc, id_keys))

    # 合并同表同操作类型的记录
    merged = merge_dml_changes(dml_changes)
    return merged


def extract_req_desc(filename):
    """从文件名提取需求描述"""
    # 格式: V202101.03.103_202511110910_T202511113698_票据标注表结构调整(标准需求).ddl.sql
    parts = filename.split('_')
    # 取最后一个有意义的部分（去掉版本号和时间戳）
    if len(parts) >= 4:
        desc_part = '_'.join(parts[3:])
        # 去掉文件扩展名
        desc_part = re.sub(r'\.(ddl|dml)\.sql$', '', desc_part, flags=re.IGNORECASE)
        return desc_part
    return filename


def extract_id_keys(sql, table, op_type):
    """从SQL中提取涉及的关键ID/KEY信息"""
    keys = []
    table_lower = table.lower()

    if op_type == 'INSERT':
        # 提取INSERT INTO后的列名列表中的ID列
        pattern = re.compile(r'(?i)INSERT\s+INTO\s+' + re.escape(table) + r'\s*\(([^)]+)\)', re.IGNORECASE)
        for m in pattern.finditer(sql):
            cols = m.group(1)
            # 找ID列
            for col in cols.split(','):
                col = col.strip().strip('"')
                if col.upper().startswith('ID') or col.upper() == 'DICT_GROUP_CODE' or col.upper() == 'KEY' or col.upper() == 'PROD_NO' or col.upper() == 'TASK_NO' or col.upper() == 'AUTH_NAME':
                    keys.append(col)
            break  # 只取第一个INSERT

    elif op_type == 'UPDATE':
        # 提取WHERE条件中的关键信息
        pattern = re.compile(r'(?i)UPDATE\s+' + re.escape(table) + r'\s+SET\s+.*?WHERE\s+(.*?)(?:;|$)', re.IGNORECASE | re.DOTALL)
        for m in pattern.finditer(sql):
            where_clause = m.group(1).strip()
            # 简化WHERE条件
            where_clause = re.sub(r'\s+', ' ', where_clause)
            if len(where_clause) > 100:
                where_clause = where_clause[:100] + '...'
            keys.append(where_clause)
            break

    if not keys:
        return ''

    return '; '.join(keys)


def merge_dml_changes(changes):
    """合并同表同操作类型的DML变更记录"""
    merged_map = {}  # (table, op_type) -> (summary_parts, id_key_parts)

    for table, op_type, summary, id_key in changes:
        key = (table.lower(), op_type)
        if key not in merged_map:
            merged_map[key] = ([], [])
        if summary and summary not in merged_map[key][0]:
            merged_map[key][0].append(summary)
        if id_key and id_key not in merged_map[key][1]:
            merged_map[key][1].append(id_key)

    result = []
    for (table, op_type), (summaries, id_keys) in merged_map.items():
        combined_summary = '; '.join(summaries) if summaries else ''
        combined_keys = '; '.join(id_keys) if id_keys else ''
        result.append((table, op_type, combined_summary, combined_keys))

    return result


# ============================================================
# Excel生成
# ============================================================

def _to_argb(hex_color):
    """将6位HEX转为openpyxl所需的aRGB格式（补FF透明度前缀）"""
    if not hex_color:
        return 'FF000000'
    color = hex_color.lstrip('#')
    # openpyxl要求8位aRGB，前两位为alpha通道
    if len(color) == 6:
        return 'FF' + color.upper()
    if len(color) == 8:
        return color.upper()
    return 'FF000000'


def build_styles(config):
    """根据配置构建Excel样式对象"""
    style_cfg = config.get('styles', {})

    # 边框样式
    border_cfg = style_cfg.get('border', {'style': 'thin', 'color': '808080'})
    border_color = _to_argb(border_cfg.get('color', '808080'))
    border = Border(
        left=Side(style=border_cfg.get('style', 'thin'), color=border_color),
        right=Side(style=border_cfg.get('style', 'thin'), color=border_color),
        top=Side(style=border_cfg.get('style', 'thin'), color=border_color),
        bottom=Side(style=border_cfg.get('style', 'thin'), color=border_color)
    )

    # 标题头样式
    header_cfg = style_cfg.get('header', {})
    header_font = Font(
        bold=header_cfg.get('bold', True),
        size=header_cfg.get('size', 12),
        color=_to_argb(header_cfg.get('font_color', 'FFFFFF'))
    )
    header_fill = PatternFill(
        start_color=_to_argb(header_cfg.get('fill_color', '1F4E78')),
        end_color=_to_argb(header_cfg.get('fill_color', '1F4E78')),
        fill_type='solid'
    )
    header_align = Alignment(
        horizontal=header_cfg.get('alignment', 'center'),
        vertical=header_cfg.get('vertical', 'center')
    )

    # 数据默认样式
    data_cfg = style_cfg.get('data', {}).get('default', {})
    data_default_font = Font(
        color=_to_argb(data_cfg.get('font_color', '000000')),
        size=data_cfg.get('size', 11)
    )

    # 数据列特殊样式
    column_styles = {}
    for col_name, col_style in style_cfg.get('data', {}).get('column_styles', {}).items():
        column_styles[col_name] = {
            'font': Font(
                bold=col_style.get('bold', False),
                color=_to_argb(col_style.get('font_color', '000000')),
                size=col_style.get('size', data_cfg.get('size', 11))
            ),
            'parentheses_color': col_style.get('parentheses_color')
        }

    return {
        'border': border,
        'header_font': header_font,
        'header_fill': header_fill,
        'header_align': header_align,
        'data_default_font': data_default_font,
        'column_styles': column_styles
    }


def _build_rich_text(value, col_style):
    """为含括号内容的列构造富文本：基础色 + 括号内浅色"""
    from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont

    text = str(value) if value is not None else ''
    paren_color = col_style.get('parentheses_color')
    base_font = col_style['font']

    # 无括号特殊处理配置 → 直接返回None
    if not paren_color:
        return None

    # 查找括号内容
    paren_match = re.search(r'([^(]*)\(([^)]*)\)(.*)', text)
    if not paren_match:
        return None

    pre, paren_content, post = paren_match.group(1), paren_match.group(2), paren_match.group(3)
    paren_argb = _to_argb(paren_color)
    base_argb = _to_argb(base_font.color.rgb if base_font.color and base_font.color.rgb else '000000')

    base_inline = InlineFont(rFont=base_font.name, sz=base_font.size, family=base_font.family, color=base_argb)
    paren_inline = InlineFont(rFont=base_font.name, sz=base_font.size, family=base_font.family, color=paren_argb)

    parts = []
    if pre:
        parts.append(TextBlock(base_inline, pre))
    parts.append(TextBlock(base_inline, '('))
    parts.append(TextBlock(paren_inline, paren_content))
    parts.append(TextBlock(base_inline, ')'))
    if post:
        parts.append(TextBlock(base_inline, post))
    return CellRichText(parts)


def generate_excel(field_changes, index_changes, dml_changes, config, output_path):
    """生成多Sheet Excel文件"""
    wb = openpyxl.Workbook()

    styles = build_styles(config)
    row_height_cfg = config.get('styles', {}).get('row_height', {'header': 25, 'data': 20})

    # Sheet1: DDL字段变更
    ddl_cfg = config['sheets']['ddl']
    ws1 = wb.active
    ws1.title = ddl_cfg['name']
    write_data_sheet(ws1, ddl_cfg['headers'], field_changes, ddl_cfg['col_widths'],
                     styles, row_height_cfg)

    # Sheet2: 索引变更
    idx_cfg = config['sheets']['index']
    ws2 = wb.create_sheet(idx_cfg['name'])
    write_data_sheet(ws2, idx_cfg['headers'], index_changes, idx_cfg['col_widths'],
                     styles, row_height_cfg)

    # Sheet3: DML配置数据变更
    dml_cfg = config['sheets']['dml']
    ws3 = wb.create_sheet(dml_cfg['name'])
    write_data_sheet(ws3, dml_cfg['headers'], dml_changes, dml_cfg['col_widths'],
                     styles, row_height_cfg)

    # 若目标文件被占用（IDE打开状态），回退到带 .new 后缀的临时文件
    final_path = output_path
    if os.path.exists(output_path):
        try:
            with open(output_path, 'a'):
                pass
        except PermissionError:
            base, ext = os.path.splitext(output_path)
            final_path = f"{base}.new{ext}"
            print(f"[WARN] 目标文件被占用，输出到临时文件: {final_path}")
    wb.save(final_path)
    print(f"Excel saved: {final_path}")
    print(f"  Sheet1 ({ddl_cfg['name']}): {len(field_changes)} rows")
    print(f"  Sheet2 ({idx_cfg['name']}): {len(index_changes)} rows")
    print(f"  Sheet3 ({dml_cfg['name']}): {len(dml_changes)} rows")


def write_data_sheet(ws, headers, data_rows, col_widths, styles, row_height_cfg):
    """写入数据到Sheet，按配置应用样式"""
    border = styles['border']
    header_font = styles['header_font']
    header_fill = styles['header_fill']
    header_align = styles['header_align']
    data_default_font = styles['data_default_font']
    column_styles = styles['column_styles']

    # 写表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = header_align
    ws.row_dimensions[1].height = row_height_cfg.get('header', 25)

    # 写数据
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            # 根据列名应用特殊样式
            col_name = headers[col_idx - 1] if col_idx - 1 < len(headers) else None
            if col_name and col_name in column_styles:
                col_style = column_styles[col_name]
                # 含括号特殊处理的列（长度）使用富文本
                rich = _build_rich_text(value, col_style)
                if rich is not None:
                    cell.value = rich
                cell.font = col_style['font']
            else:
                cell.font = data_default_font
        ws.row_dimensions[row_idx].height = row_height_cfg.get('data', 20)

    # 设置列宽
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width


# ============================================================
# 主流程
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='SQL增量脚本变更分析器')
    parser.add_argument('--config', '-c', help='配置文件路径', default=None)
    parser.add_argument('--sql-dir', '-s', help='SQL脚本目录（覆盖配置）', default=None)
    parser.add_argument('--output', '-o', help='输出文件路径（覆盖配置）', default=None)
    parser.add_argument('--project-root', '-p', help='项目根目录', default='.')
    args = parser.parse_args()

    # 加载配置
    config = load_config(args.config)
    config['_project_root'] = os.path.abspath(args.project_root)

    # 命令行参数覆盖
    sql_dir = args.sql_dir or resolve_path(config, 'sql_dir')
    output_path = args.output or resolve_path(config, 'output_dir')
    if os.path.isdir(output_path):
        output_path = os.path.join(output_path, config['output_filename_pattern'])

    print(f"SQL dir: {sql_dir}")
    print(f"Output: {output_path}")

    # 扫描SQL文件
    ddl_files, dml_files = scan_sql_files(sql_dir, config)

    # 解析DDL
    field_changes, index_changes = parse_ddl_files(ddl_files, config)

    # 解析DML
    dml_changes = parse_dml_files(dml_files, config)

    # 生成Excel
    generate_excel(field_changes, index_changes, dml_changes, config, output_path)

    print("\n=== Summary ===")
    print(f"DDL field changes: {len(field_changes)}")
    print(f"Index changes: {len(index_changes)}")
    print(f"DML config changes: {len(dml_changes)}")

    # 统计涉及表
    tables = set()
    for fc in field_changes:
        tables.add(fc[0].lower())
    for ic in index_changes:
        tables.add(ic[0].lower())
    for dc in dml_changes:
        tables.add(dc[0].lower())
    print(f"Total tables affected: {len(tables)}")


if __name__ == '__main__':
    main()
